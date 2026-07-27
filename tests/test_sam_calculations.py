"""Core SAM calculations vs. hand-verified expected values.

These are the "tool output vs. manually verified examples" tests: the input
series and the manual arithmetic are documented in tests/conftest.py.
"""
from __future__ import annotations

from openpyxl import load_workbook

from app.modules.sam_report import excel_writer, parser, service


def test_max_rolling_avg_smooths_a_spike():
    # window 3, shrinks at the series start.
    # [0,0,12,12,12,0,30] -> avgs 0,0,4,8,12,8,14 -> max 14
    assert parser._max_rolling_avg([0, 0, 12, 12, 12, 0, 30], 3) == 14.0
    # [10,20,30,40,5,5,5] -> avgs 10,15,20,30,25,16.67,5 -> max 30 (raw max is 40)
    assert parser._max_rolling_avg([10, 20, 30, 40, 5, 5, 5], 3) == 30.0


def test_extract_runs_matches_hand_computed_values(sam_workbook):
    runs, warnings = parser.extract_runs(sam_workbook)

    assert warnings == []                    # a clean, well-formed workbook
    assert [r.year for r in runs] == [2001, 2002]

    r1, r2 = runs
    assert (r1.max_isc_a, r1.max_voc_v, r1.max_isc_rolling_avg_a) == (30.0, 1379.54, 14.0)
    assert (r2.max_isc_a, r2.max_voc_v, r2.max_isc_rolling_avg_a) == (40.0, 1441.93, 30.0)


def test_build_table_rows_and_maximum_footer(sam_workbook):
    runs, _ = parser.extract_runs(sam_workbook)
    rows = service._build_table(runs)
    year, isc, voc, roll = excel_writer.TABLE_COLUMNS

    assert [r.cells[year] for r in rows] == [2001, 2002, "Maximum"]
    # per-run cells are 2-dp strings
    assert rows[0].cells[roll] == "14.00"
    assert rows[1].cells[isc] == "40.00"
    # the Maximum footer is the column-wise max across runs
    mx = rows[-1].cells
    assert (mx[isc], mx[voc], mx[roll]) == ("40.00", "1441.93", "30.00")


def test_output_workbook_layout_and_values(sam_workbook, tmp_path):
    runs, _ = parser.extract_runs(sam_workbook)
    out = excel_writer.write_output_workbook(sam_workbook, runs, tmp_path / "out.xlsx")

    wb = load_workbook(out)
    assert excel_writer.OUTPUT_SHEET_NAME in wb.sheetnames
    ws = wb[excel_writer.OUTPUT_SHEET_NAME]

    # header anchored at B4
    assert [ws.cell(4, c).value for c in (2, 3, 4, 5)] == excel_writer.TABLE_COLUMNS
    # first data row (B5..E5): year, then FULL-precision numbers (not the 2-dp display)
    assert [ws.cell(5, c).value for c in (2, 3, 4, 5)] == [2001, 30.0, 1379.54, 14.0]
    # Maximum footer at row 7 (2 runs -> rows 5,6, footer 7)
    assert ws.cell(7, 2).value == "Maximum"
    assert [ws.cell(7, c).value for c in (3, 4, 5)] == [40.0, 1441.93, 30.0]
