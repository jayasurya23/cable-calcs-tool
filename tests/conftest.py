"""Shared test fixtures for the SAM-report pipeline.

The fixtures build small, *format-correct* inputs with values chosen so every
expected output can be worked out by hand (see the comments below). This avoids
committing the 2.8 MB reference workbook while still exercising the real parser.

Hand-computed reference values (rolling window = 3 samples for this short series,
because steps_per_hour = max(1, round(len/8760)) = 1):

  Run 1 (weather year 2001)
    Isc = [0, 0, 12, 12, 12, 0, 30]
      max Isc               = 30
      3-sample rolling avgs = 0, 0, 4, 8, 12, 8, 14   -> max = 14
    Voc = [800, 810, 805, 1379.54, 800, 800, 800]     -> max Voc = 1379.54

  Run 2 (weather year 2002)
    Isc = [10, 20, 30, 40, 5, 5, 5]
      max Isc               = 40
      3-sample rolling avgs = 10, 15, 20, 30, 25, 16.67, 5  -> max = 30
    Voc = [900, 1441.93, 900, 900, 900, 900, 900]     -> max Voc = 1441.93

  "Maximum" row (column-wise max across runs):
    max Isc = 40 · max Voc = 1441.93 · max rolling = 30
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook

# ── Known input series (hand-verified expectations live in the module docstring)
RUN1_ISC = [0, 0, 12, 12, 12, 0, 30]
RUN1_VOC = [800, 810, 805, 1379.54, 800, 800, 800]
RUN2_ISC = [10, 20, 30, 40, 5, 5, 5]
RUN2_VOC = [900, 1441.93, 900, 900, 900, 900, 900]
# The year is taken from the LAST 19xx/20xx in the weather-file name.
RUN1_WEATHER = "phoenix_33.4500_-112.0700_nsrdb_60_2001.csv"   # -> 2001
RUN2_WEATHER = "phoenix_33.4500_-112.0700_nsrdb_60_2002.csv"   # -> 2002

# SAM sheet titles Excel-truncated to <= 31 chars, still carrying the keywords
# the parser fuzzy-matches on ("open circuit" / "short circuit" / "solar resource").
VOC_SHEET = "Subarray 1 Open circuit DC vol"
ISC_SHEET = "Subarray 1 String short circuit"
WX_SHEET = "IN_Solar resource library curr"


def _hourly_sheet(wb: Workbook, title: str, run1: list, run2: list):
    ws = wb.create_sheet(title)
    ws.append(["Hour", "Run 1", "Run 2"])
    for hour, (a, b) in enumerate(zip(run1, run2)):
        ws.append([hour, a, b])
    return ws


@pytest.fixture
def sam_workbook(tmp_path):
    """A minimal, format-correct SAM parametric-runs workbook (2 runs)."""
    wb = Workbook()
    del wb["Sheet"]                                   # drop the default empty sheet
    _hourly_sheet(wb, VOC_SHEET, RUN1_VOC, RUN2_VOC)
    _hourly_sheet(wb, ISC_SHEET, RUN1_ISC, RUN2_ISC)
    wx = wb.create_sheet(WX_SHEET)
    wx.append(["Hour", "Run 1", "Run 2"])
    wx.append(["", RUN1_WEATHER, RUN2_WEATHER])       # one row of weather-file names
    path = tmp_path / "sam_runs.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def pysam_inputs() -> dict:
    """A minimal pvsamv1-style inputs dict. SAM stores every value as a string,
    picks equipment by numeric selector, and rarely carries a product name."""
    return {
        "module_model": "1",                          # CEC database
        "cec_v_mp_ref": "38.5", "cec_i_mp_ref": "13.9",
        "cec_v_oc_ref": "45.8", "cec_i_sc_ref": "14.6", "cec_n_s": "144",
        "inverter_model": "0",                        # CEC database (Sandia)
        "inv_snl_paco": "125000", "inv_snl_pdco": "127000",
        "inv_snl_vdco": "600", "inv_snl_vdcmax": "1000", "inverter_count": "16",
        "mppt_low_inverter": "500", "mppt_hi_inverter": "800", "inv_num_mppt": "1",
        "subarray1_gcr": "0.36", "subarray1_modules_per_string": "26",
        "system_capacity": "2600",                    # kW DC
        "solar_resource_file": "C:/wx/phoenix_33.4500_-112.0700_nsrdb_60_2001.csv",
        "use_wf_albedo": "0", "albedo": ["0.25"] * 12,
    }
