"""Parser robustness: value coercion, fuzzy sheet matching, year logic,
graceful degradation, and pysam equipment extraction."""
from __future__ import annotations

from openpyxl import Workbook

from app.modules.sam_report import parser


def test_to_finite_float_coerces_text_and_junk():
    assert parser._to_finite_float("12.5") == 12.5          # SAM writes numbers as text
    assert parser._to_finite_float(7) == 7.0
    assert parser._to_finite_float(None) == 0.0             # blank cell
    assert parser._to_finite_float("not a number") == 0.0
    assert parser._to_finite_float(float("nan")) == 0.0     # never reaches Isc/Voc cells
    assert parser._to_finite_float(float("inf")) == 0.0


def test_year_from_weather_file_takes_last_match():
    assert parser._year_from_weather_file("x_60_2001.csv") == 2001
    assert parser._year_from_weather_file("nsrdb_2019_60_2024.csv") == 2024   # LAST year wins
    assert parser._year_from_weather_file("no_year_here.csv") is None


def test_find_sheet_is_fuzzy_and_keyword_based():
    names = ["Metadata", "Subarray 1 Open circuit DC vol", "Subarray 1 String short circuit"]
    assert parser._find_sheet(names, parser.VOC_SHEET_KEYWORDS) == "Subarray 1 Open circuit DC vol"
    assert parser._find_sheet(names, parser.ISC_SHEET_KEYWORDS) == "Subarray 1 String short circuit"
    assert parser._find_sheet(names, parser.WEATHER_SHEET_KEYWORDS) is None


def test_non_sam_workbook_degrades_with_warning(tmp_path):
    wb = Workbook()
    wb.active.title = "RandomData"
    wb.active.append(["a", "b", "c"])
    path = tmp_path / "notsam.xlsx"
    wb.save(path)

    runs, warnings = parser.extract_runs(path)
    assert runs == []
    assert warnings and "does not look like a SAM" in warnings[0]


def test_missing_weather_sheet_falls_back_to_1998_consecutive(tmp_path):
    wb = Workbook()
    del wb["Sheet"]
    for title, v in [("Subarray 1 Open circuit DC vol", (800, 810)),
                     ("Subarray 1 String short circuit", (10, 20))]:
        ws = wb.create_sheet(title)
        ws.append(["Hour", "Run 1", "Run 2"])
        ws.append([0, v[0], v[0]])
        ws.append([1, v[1], v[1]])
    path = tmp_path / "noweather.xlsx"
    wb.save(path)

    runs, warnings = parser.extract_runs(path)
    assert [r.year for r in runs] == [1998, 1999]           # Run 1 = 1998, consecutive
    assert any("Run 1 = 1998" in w for w in warnings)


def test_extract_equipment_reads_selectors_and_specs(pysam_inputs):
    eq = parser.extract_equipment(pysam_inputs)

    assert eq.module_model_type == "CEC Performance Model (module database)"
    assert eq.module_specs["Pmax"] == "535.15 W"            # 38.5 * 13.9, shown exactly
    assert eq.module_specs["Voc"] == "45.8 V"
    assert eq.module_specs["Cells"] == "144"

    assert eq.inverter_model_type == "Inverter CEC Database (Sandia)"
    assert eq.inverter_specs["Pac"] == "125.0 kW"           # 125000 W / 1000
    assert eq.inverter_specs["MPPT range"] == "500–800 V"
    assert eq.inverter_specs["Qty"] == "16"

    # this export has no product-name string -> a note explains the fallback
    assert eq.module_model is None
    assert eq.notes
