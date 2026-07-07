"""
Writes the standardized "Output" sheet into a copy of the uploaded SAM workbook.

Design replicated from the reference sample ("SAM excel output.xlsx"):
  * Table anchored at B4; columns: Year | Max SAM Isc (A) | Max SAM Voc (V) |
    Max SAM Isc Rolling Average (A); one row per run-year; "Maximum" footer row.
  * Header + footer rows: solid fill, theme color 3 @ 75% tint ("Text 2,
    Lighter 75%" — renders from the workbook's own theme, so it matches SAM's
    export exactly), thin borders.
  * All cells centered, Aptos Narrow 11; numeric cells formatted "0.00"
    (full precision kept underneath, like the sample).
  * Column widths taken from the sample (B..E = 17.4/19.1/21.4/33.6).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import SamRun

OUTPUT_SHEET_NAME = "Output"

TABLE_COLUMNS = [
    "Year",
    "Max SAM Isc (A)",
    "Max SAM Voc (V)",
    "Max SAM Isc Rolling Average (A)",
]

_ANCHOR_ROW = 4      # header row (matches the sample)
_ANCHOR_COL = 2      # column B
_COL_WIDTHS = [17.4, 19.1, 21.4, 33.6]

_FONT = Font(name="Aptos Narrow", size=11)
_CENTER = Alignment(horizontal="center", vertical="center")
_HEADER_FILL = PatternFill(
    patternType="solid", fgColor=Color(theme=3, tint=0.749992370372631)
)
_THIN = Side(style="thin")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NUMFMT = "0.00"


def _write_row(ws, row_idx: int, values: list, filled: bool) -> None:
    for j, val in enumerate(values):
        cell = ws.cell(row=row_idx, column=_ANCHOR_COL + j, value=val)
        cell.font = _FONT
        cell.alignment = _CENTER
        cell.border = _BOX
        if filled:
            cell.fill = _HEADER_FILL
        if j > 0 and isinstance(val, (int, float)):
            cell.number_format = _NUMFMT


def write_output_workbook(
    src_path: str | Path, runs: list[SamRun], dest_path: str | Path
) -> Path:
    """Copy the uploaded workbook and add/replace its standardized Output sheet.

    Loads with data_only=False so any formulas elsewhere in the engineer's
    workbook survive the round trip untouched.
    """
    src_path, dest_path = Path(src_path), Path(dest_path)
    wb = load_workbook(filename=str(src_path), data_only=False)
    try:
        if OUTPUT_SHEET_NAME in wb.sheetnames:
            del wb[OUTPUT_SHEET_NAME]
        ws = wb.create_sheet(OUTPUT_SHEET_NAME)

        for j, width in enumerate(_COL_WIDTHS):
            ws.column_dimensions[get_column_letter(_ANCHOR_COL + j)].width = width

        _write_row(ws, _ANCHOR_ROW, TABLE_COLUMNS, filled=True)
        r = _ANCHOR_ROW + 1
        for run in runs:
            _write_row(
                ws, r,
                [run.year, run.max_isc_a, run.max_voc_v, run.max_isc_rolling_avg_a],
                filled=False,
            )
            r += 1
        _write_row(
            ws, r,
            [
                "Maximum",
                max((x.max_isc_a for x in runs), default=0.0),
                max((x.max_voc_v for x in runs), default=0.0),
                max((x.max_isc_rolling_avg_a for x in runs), default=0.0),
            ],
            filled=True,
        )

        wb.save(str(dest_path))
    finally:
        wb.close()
    return dest_path
