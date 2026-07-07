"""
Parsing for the SAM parametric-runs workbook (+ optional pysam JSON).

Expected workbook layout (from SAM's parametric export, verified against the
sample "SAM excel output.xlsx"):

  * A "Subarray 1 Open circuit DC volt…" sheet — col A = Hour (0..8759),
    then one column per run ("Run 1", "Run 2", …) of hourly subarray Voc.
  * A "Subarray 1 String short circuit…" sheet — same shape, hourly string Isc.
  * An "IN_Solar resource library curre…" sheet — one weather-file name per run;
    names end in the weather year (…_60_1998). Run 1 starts at 1998 and runs
    are consecutive years, so the year also falls back to 1998 + index.

Sheet names are matched fuzzily (Excel truncates titles at 31 chars, and SAM
may vary suffixes) — see the *_KEYWORDS constants.
"""
from __future__ import annotations

import json
import math
import re
from pathlib import Path

from openpyxl import load_workbook

from .models import Equipment, SamRun, SheetPreview

_MAX_SAMPLE_ROWS = 15
_MAX_COLS = 60

# Fuzzy sheet matching: every keyword must appear in the lowercased title.
VOC_SHEET_KEYWORDS = ("open circuit",)
ISC_SHEET_KEYWORDS = ("short circuit",)
WEATHER_SHEET_KEYWORDS = ("solar resource",)

FIRST_RUN_YEAR_DEFAULT = 1998
ROLLING_WINDOW_HOURS = 3

_YEAR_RE = re.compile(r"(19|20)\d{2}")


def preview_workbook(path: str | Path) -> list[SheetPreview]:
    """Open a workbook and return a light preview of every sheet.

    Read-only + values-only so large SAM exports load fast and we never trip on
    formulas. This works regardless of the run schema.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    previews: list[SheetPreview] = []
    try:
        for ws in wb.worksheets:
            n_rows = ws.max_row or 0
            n_cols = min(ws.max_column or 0, _MAX_COLS)

            rows_iter = ws.iter_rows(values_only=True)
            header: list[str] = []
            sample_rows: list[list[str]] = []
            for i, row in enumerate(rows_iter):
                cells = ["" if c is None else str(c) for c in row[:_MAX_COLS]]
                if i == 0:
                    header = cells
                elif i <= _MAX_SAMPLE_ROWS:
                    sample_rows.append(cells)
                else:
                    break

            previews.append(
                SheetPreview(
                    name=ws.title,
                    n_rows=n_rows,
                    n_cols=n_cols,
                    header=header,
                    sample_rows=sample_rows,
                )
            )
    finally:
        wb.close()
    return previews


def parse_pysam_json(path: str | Path) -> dict:
    """Load a pysam case-inputs JSON (SAM: Generate code -> JSON for inputs)."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ─── pysam equipment extraction ──────────────────────────────────────────────
#
# A pvsamv1 inputs JSON is flat and stores every value as a string. The module
# and inverter are chosen by numeric selector keys (module_model / inverter_model)
# and defined by a family of parameters (cec_*, inv_snl_*, …). There is usually
# NO product-name string. So we map the selector to a model-type label and pull
# the defining nameplate specs from the active parameter family. If the JSON
# *does* carry a real product name (some exports do), we surface that too.
#
# We still flatten nested dicts first, so a grouped/nested JSON also works.

_MODULE_NAME_KEYS = (
    "module_name", "module_model_name", "cec_module_name",
    "solar_module_name", "mod_name", "modulename",
)
_INVERTER_NAME_KEYS = (
    "inverter_name", "inv_name", "inverter_model_name",
    "inv_snl_name", "invertername",
)

# pvsamv1 module_model selector -> label.
_MODULE_MODEL_TYPES = {
    "0": "Simple Efficiency Module Model",
    "1": "CEC Performance Model (module database)",
    "2": "CEC Performance Model (user-entered specs)",
    "3": "Sandia PV Array Performance Model (database)",
    "4": "IEC 61853 Single Diode Model",
    "5": "Mermoud-Lejeune Single Diode Model",
}
# Per module_model: (Vmp, Imp, Voc, Isc, N_cells) parameter keys.
_MODULE_SPEC_KEYS = {
    "1": ("cec_v_mp_ref", "cec_i_mp_ref", "cec_v_oc_ref", "cec_i_sc_ref", "cec_n_s"),
    "2": ("cec_v_mp_ref", "cec_i_mp_ref", "cec_v_oc_ref", "cec_i_sc_ref", "cec_n_s"),
    "3": ("snl_vmpo", "snl_impo", "snl_voco", "snl_isco", "snl_series_cells"),
    "4": ("sd11par_Vmp0", "sd11par_Imp0", "sd11par_Voc0", "sd11par_Isc0", "sd11par_Nser"),
    "5": ("mlm_V_mp_ref", "mlm_I_mp_ref", "mlm_V_oc_ref", "mlm_I_sc_ref", "mlm_N_series"),
}

# pvsamv1 inverter_model selector -> label.
_INVERTER_MODEL_TYPES = {
    "0": "Inverter CEC Database (Sandia)",
    "1": "Inverter Datasheet",
    "2": "Inverter Part-Load Curve",
    "3": "Inverter CEC Coefficient Generator",
}
# Per inverter_model: (Paco, Pdco, Vdco, Vdcmax) parameter keys (Pdco may be None).
_INVERTER_SPEC_KEYS = {
    "0": ("inv_snl_paco", "inv_snl_pdco", "inv_snl_vdco", "inv_snl_vdcmax"),
    "1": ("inv_ds_paco", None, "inv_ds_vdco", "inv_ds_vdcmax"),
    "2": ("inv_pd_paco", "inv_pd_pdco", "inv_pd_vdco", "inv_pd_vdcmax"),
    "3": ("inv_cec_cg_paco", "inv_cec_cg_pdco", "inv_cec_cg_vdco", "inv_cec_cg_vdcmax"),
}


def _flatten_json(data, parent: str = "", sep: str = ".") -> dict[str, object]:
    """Flatten nested dicts to dotted leaf keys. Lists collapse to their value."""
    out: dict[str, object] = {}
    if isinstance(data, dict):
        for k, v in data.items():
            key = f"{parent}{sep}{k}" if parent else str(k)
            out.update(_flatten_json(v, key, sep))
    else:
        out[parent] = data
    return out


def _leaf(key: str) -> str:
    return key.split(".")[-1].lower()


def _looks_like_name(v) -> bool:
    """A product name has letters and isn't just a number (SAM stores '2' as a str)."""
    if not isinstance(v, str):
        return False
    s = v.strip()
    return len(s) > 2 and any(c.isalpha() for c in s) and not s.replace(".", "", 1).isdigit()


def _find_named(flat: dict[str, object], candidates: tuple[str, ...]) -> str | None:
    """Exact leaf-key match against name-bearing keys (value must look like a name)."""
    for cand in candidates:
        for k, v in flat.items():
            if _leaf(k) == cand and _looks_like_name(v):
                return v.strip()
    return None


def _num(flat: dict[str, object], key: str | None):
    """Parse a numeric value (SAM stores numbers as strings). None if absent/bad."""
    if not key:
        return None
    for k, v in flat.items():
        if _leaf(k) == key.lower():
            try:
                return float(v)
            except (TypeError, ValueError):
                return None
    return None


def _fmt(v, unit: str, dp: int = 0) -> str:
    return f"{v:.{dp}f} {unit}" if dp else f"{v:g} {unit}"


def _module_specs(flat, model: str) -> dict[str, str]:
    keys = _MODULE_SPEC_KEYS.get(model)
    if not keys:
        return {}
    vmp, imp, voc, isc, ns = (_num(flat, k) for k in keys)
    specs: dict[str, str] = {}
    if vmp is not None and imp is not None:
        specs["Pmax"] = _fmt(vmp * imp, "W", 0)
    if vmp is not None:
        specs["Vmp"] = _fmt(vmp, "V")
    if imp is not None:
        specs["Imp"] = _fmt(imp, "A")
    if voc is not None:
        specs["Voc"] = _fmt(voc, "V")
    if isc is not None:
        specs["Isc"] = _fmt(isc, "A")
    if ns is not None:
        specs["Cells"] = f"{int(ns)}"
    return specs


def _inverter_specs(flat, model: str) -> dict[str, str]:
    keys = _INVERTER_SPEC_KEYS.get(model)
    specs: dict[str, str] = {}
    if keys:
        paco, pdco, vdco, vdcmax = (_num(flat, k) for k in keys)
        if paco is not None:
            specs["Pac"] = _fmt(paco / 1000.0, "kW", 1)
        if pdco is not None:
            specs["Pdc"] = _fmt(pdco / 1000.0, "kW", 1)
        if vdcmax is not None:
            specs["Vdc max"] = _fmt(vdcmax, "V")
        if vdco is not None:
            specs["Vdco"] = _fmt(vdco, "V")
    lo, hi = _num(flat, "mppt_low_inverter"), _num(flat, "mppt_hi_inverter")
    if lo is not None and hi is not None:
        specs["MPPT range"] = f"{lo:g}–{hi:g} V"
    nmppt = _num(flat, "inv_num_mppt")
    if nmppt is not None:
        specs["MPPTs"] = f"{int(nmppt)}"
    qty = _num(flat, "inverter_count")
    if qty is not None:
        specs["Qty"] = f"{int(qty)}"
    return specs


def extract_equipment(data: dict) -> Equipment:
    """Extract module/inverter model type, nameplate specs, and any name strings."""
    flat = _flatten_json(data)

    module_name = _find_named(flat, _MODULE_NAME_KEYS)
    inverter_name = _find_named(flat, _INVERTER_NAME_KEYS)

    mod_sel = next((str(v).strip() for k, v in flat.items()
                    if _leaf(k) == "module_model"), None)
    inv_sel = next((str(v).strip() for k, v in flat.items()
                    if _leaf(k) == "inverter_model"), None)

    module_type = _MODULE_MODEL_TYPES.get(mod_sel) if mod_sel is not None else None
    inverter_type = _INVERTER_MODEL_TYPES.get(inv_sel) if inv_sel is not None else None

    module_specs = _module_specs(flat, mod_sel) if mod_sel is not None else {}
    inverter_specs = _inverter_specs(flat, inv_sel) if inv_sel is not None else {}

    notes: list[str] = []
    if (module_type or inverter_type) and not (module_name or inverter_name):
        notes.append(
            "This pysam file defines equipment by parameters, not product names — "
            "showing the SAM model type and nameplate specs."
        )

    return Equipment(
        module_model=module_name,
        inverter_model=inverter_name,
        module_model_type=module_type,
        inverter_model_type=inverter_type,
        module_specs=module_specs,
        inverter_specs=inverter_specs,
        notes=notes,
    )


# ─── SAM run extraction ──────────────────────────────────────────────────────

def _find_sheet(sheetnames: list[str], keywords: tuple[str, ...]) -> str | None:
    for name in sheetnames:
        low = name.lower()
        if all(k in low for k in keywords):
            return name
    return None


def _read_run_series(ws) -> tuple[list[str], list[list[float]]]:
    """Read a value sheet: header row ('Hour', 'Run 1', ...), then hourly rows.

    Returns (run_ids, series) where series[i] is the full hourly list for run i.
    SAM writes numbers as text sometimes, so every cell goes through float().
    """
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    run_ids = [h for h in header[1:] if h]
    series: list[list[float]] = [[] for _ in run_ids]
    for row in rows:
        for i in range(len(run_ids)):
            series[i].append(_to_finite_float(row[i + 1]))
    return run_ids, series


def _to_finite_float(v) -> float:
    """Coerce a cell to a finite float; blank / non-numeric / NaN / inf -> 0.0.

    SAM sometimes writes numbers as text, and a malformed/hand-edited export can
    carry 'nan'/'inf' or junk — those must never reach the report as literal
    'nan'/'inf' in the Isc/Voc cells that feed NEC sizing.
    """
    if v is None:
        return 0.0
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0.0
    return f if math.isfinite(f) else 0.0


def _read_weather_files(ws) -> dict[str, str]:
    """Read the solar-resource-library sheet: {'Run 1': '<weather file name>', ...}."""
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    try:
        values = next(rows)
    except StopIteration:
        return {}
    out: dict[str, str] = {}
    for i, run_id in enumerate(header[1:], start=1):
        if run_id and i < len(values) and values[i] is not None:
            out[run_id] = str(values[i])
    return out


def _year_from_weather_file(name: str) -> int | None:
    """Weather file names end in the year (…nsrdb-GOES…_60_1998) — take the last one."""
    last = None
    for m in _YEAR_RE.finditer(name):
        last = m.group(0)
    return int(last) if last else None


def _max_rolling_avg(vals: list[float], window: int) -> float:
    """Max of the trailing rolling average (window shrinks at the series start)."""
    best = float("-inf")
    running = 0.0
    for t, v in enumerate(vals):
        running += v
        if t >= window:
            running -= vals[t - window]
        n = min(t + 1, window)
        avg = running / n
        if avg > best:
            best = avg
    return best


def extract_runs(workbook_path: str | Path) -> tuple[list[SamRun], list[str]]:
    """Extract per-run stats from a SAM runs workbook.

    Returns (runs, warnings). Empty runs + a warning if the expected SAM value
    sheets are missing (the caller falls back to the raw preview).
    """
    warnings: list[str] = []
    wb = load_workbook(filename=str(workbook_path), read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        voc_name = _find_sheet(names, VOC_SHEET_KEYWORDS)
        isc_name = _find_sheet(names, ISC_SHEET_KEYWORDS)
        if not voc_name or not isc_name:
            warnings.append(
                "Workbook does not look like a SAM runs export — expected "
                "'…Open circuit DC volt…' and '…String short circuit…' sheets."
            )
            return [], warnings

        isc_run_ids, isc_series = _read_run_series(wb[isc_name])
        voc_run_ids, voc_series = _read_run_series(wb[voc_name])
        if isc_run_ids != voc_run_ids:
            warnings.append(
                f"Run columns differ between '{isc_name}' ({len(isc_run_ids)}) and "
                f"'{voc_name}' ({len(voc_run_ids)}); using the short-circuit sheet's runs."
            )

        weather_files: dict[str, str] = {}
        weather_name = _find_sheet(names, WEATHER_SHEET_KEYWORDS)
        if weather_name:
            weather_files = _read_weather_files(wb[weather_name])
        else:
            warnings.append(
                "No solar-resource-library sheet found; assuming Run 1 = "
                f"{FIRST_RUN_YEAR_DEFAULT} with consecutive years."
            )

        # 3-hr rolling window in samples: hourly data -> 3; sub-hourly scales up.
        n = len(isc_series[0]) if isc_series else 0
        steps_per_hour = max(1, round(n / 8760)) if n else 1
        window = ROLLING_WINDOW_HOURS * steps_per_hour

        voc_by_run = dict(zip(voc_run_ids, voc_series))
        runs: list[SamRun] = []
        for i, run_id in enumerate(isc_run_ids):
            isc = isc_series[i]
            voc = voc_by_run.get(run_id, [])
            wf = weather_files.get(run_id)
            year = (_year_from_weather_file(wf) if wf else None) or (
                FIRST_RUN_YEAR_DEFAULT + i
            )
            runs.append(
                SamRun(
                    run_id=run_id,
                    year=year,
                    weather_file=wf,
                    max_isc_a=max(isc) if isc else 0.0,
                    max_voc_v=max(voc) if voc else 0.0,
                    max_isc_rolling_avg_a=_max_rolling_avg(isc, window) if isc else 0.0,
                )
            )
        return runs, warnings
    finally:
        wb.close()
